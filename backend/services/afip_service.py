"""
Servicio AFIP - Mis Comprobantes (fes.afip.gob.ar/mcmp).
Flujo verificado:
  1. Login AFIP
  2. Click "Mis Comprobantes" en portal → nueva pestaña en fes.afip.gob.ar
  3. Click en representado (link con nombre)
  4. Click #btnEmitidos / #btnRecibidos
  5. Llenar #fechaEmision con "DD/MM/YYYY - DD/MM/YYYY"
  6. Click #buscarComprobantes
  7. Click CSV (class 'sinbor')
"""
import os
import asyncio
import calendar
import zipfile
import csv
import io
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv

load_dotenv()

STORAGE_PATH = os.getenv("STORAGE_PATH", "./storage")
AFIP_LOGIN_URL = "https://auth.afip.gob.ar/contribuyente_/login.xhtml"

_executor = ThreadPoolExecutor(max_workers=2)


def _csv_a_xlsx(csv_bytes: bytes, xlsx_path: str):
    """Convierte el CSV de AFIP (sep=; decimal=,) a .xlsx con celdas numéricas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers as xl_numbers

    texto = csv_bytes.decode("utf-8-sig")  # utf-8-sig quita BOM si existe
    reader = csv.reader(io.StringIO(texto), delimiter=";")
    filas = list(reader)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Estilo encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")

    for i, fila in enumerate(filas):
        for j, valor in enumerate(fila):
            valor = valor.strip().strip('"')
            celda = ws.cell(row=i + 1, column=j + 1)

            if i == 0:
                celda.value = valor
                celda.font = header_font
                celda.fill = header_fill
                celda.alignment = Alignment(horizontal="center")
            else:
                # Intentar convertir a número (formato argentino: coma decimal)
                try:
                    num = float(valor.replace(".", "").replace(",", "."))
                    celda.value = num
                    # Si tiene decimales, formato numérico con 2 decimales
                    if "," in valor:
                        celda.number_format = '#,##0.00'
                    else:
                        celda.number_format = '#,##0'
                except ValueError:
                    celda.value = valor

    # Autoajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(xlsx_path)


def _strip_cuit(cuit: str) -> str:
    return cuit.replace("-", "").replace(" ", "")


def _periodo_a_fechas(periodo_desde: str, periodo_hasta: str) -> tuple:
    """
    Convierte fecha a DD/MM/YYYY para el daterangepicker de AFIP.
    Acepta YYYY-MM-DD (fecha puntual) o YYYY-MM (mes completo).
    """
    partes_d = periodo_desde.split("-")
    partes_h = periodo_hasta.split("-")

    if len(partes_d) == 3:
        # YYYY-MM-DD
        anio_d, mes_d, dia_d = partes_d
    else:
        anio_d, mes_d, dia_d = partes_d[0], partes_d[1], "01"

    if len(partes_h) == 3:
        anio_h, mes_h, dia_h = partes_h
    else:
        anio_h, mes_h = partes_h[0], partes_h[1]
        dia_h = str(calendar.monthrange(int(anio_h), int(mes_h))[1])

    fecha_desde = f"{int(dia_d):02d}/{int(mes_d):02d}/{anio_d}"
    fecha_hasta = f"{int(dia_h):02d}/{int(mes_h):02d}/{anio_h}"
    return fecha_desde, fecha_hasta


def _descargar_todos_sync(
    cuit: str,
    password: str,
    representado: str,
    periodo_desde: str,
    periodo_hasta: str,
    tipos: list,
    cliente_cuit: str,
    download_dir: str,
) -> dict:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

    Path(download_dir).mkdir(parents=True, exist_ok=True)
    cuit_digits = _strip_cuit(cuit)
    resultados = {tipo: (None, "No ejecutado") for tipo in tipos}

    fecha_desde, fecha_hasta = _periodo_a_fechas(periodo_desde, periodo_hasta)
    rango_fecha = f"{fecha_desde} - {fecha_hasta}"

    print(f"[AFIP] CUIT={cuit_digits} representado='{representado}' fechas='{rango_fecha}'")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        try:
            # ── 1. LOGIN ──────────────────────────────────────────────
            print("[AFIP] Login...")
            page.goto(AFIP_LOGIN_URL, timeout=30000)
            page.wait_for_load_state("networkidle", timeout=15000)
            page.fill('[id="F1:username"]', cuit_digits)
            page.click('[id="F1:btnSiguiente"]')
            page.wait_for_selector('[id="F1:password"]', timeout=15000)
            page.fill('[id="F1:password"]', password)
            page.click('[id="F1:btnIngresar"]')
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(5000)
            print(f"[AFIP] Login OK. URL: {page.url}")

            if "login" in page.url or "error" in page.url.lower():
                for tipo in tipos:
                    resultados[tipo] = (None, "Credenciales AFIP inválidas")
                return resultados

            # ── 2. CLICK EN MIS COMPROBANTES ──────────────────────────
            print("[AFIP] Buscando 'Mis Comprobantes' en portal...")
            mc_link = None
            for el in page.locator("a").all():
                try:
                    if el.inner_text(timeout=300).strip() == "Mis Comprobantes":
                        mc_link = el
                        break
                except Exception:
                    pass

            if not mc_link:
                for tipo in tipos:
                    resultados[tipo] = (None, "No se encontró 'Mis Comprobantes' en el portal AFIP")
                return resultados

            with context.expect_page() as new_page_info:
                mc_link.click()
            mc = new_page_info.value
            mc.wait_for_load_state("networkidle", timeout=20000)
            mc.wait_for_timeout(3000)
            print(f"[AFIP] Mis Comprobantes abierto. URL: {mc.url}")

            # ── 3. SELECCIONAR REPRESENTADO ───────────────────────────
            # Detectar si ya estamos en el menú (persona física = único representado)
            # El indicador más confiable es la existencia de #btnEmitidos en la página
            try:
                ya_en_menu = mc.locator("#btnEmitidos").count() > 0
            except Exception:
                ya_en_menu = False

            if not ya_en_menu:
                # Segunda verificación por URL
                ya_en_menu = "menuPrincipal" in mc.url or "comprobantesEmitidos" in mc.url

            if ya_en_menu:
                print(f"[AFIP] Portal ya en menú (persona física sin selección de representado). URL: {mc.url}")
            else:
                print(f"[AFIP] Seleccionando representado: '{representado}'")

                palabras_clave = [p.upper() for p in representado.split() if len(p) > 3]
                print(f"[AFIP] Palabras clave: {palabras_clave}")

                rep_link = None
                todos_links_texto = []
                for el in mc.locator("a").all():
                    try:
                        text = el.inner_text(timeout=300).strip()
                        if not text:
                            continue
                        todos_links_texto.append(text[:80])
                        text_upper = text.upper()
                        if representado.upper() in text_upper:
                            rep_link = el
                            print(f"[AFIP] Match exacto: '{text[:60]}'")
                            break
                        if palabras_clave and all(p in text_upper for p in palabras_clave):
                            rep_link = el
                            print(f"[AFIP] Match por palabras clave: '{text[:60]}'")
                            break
                    except Exception:
                        pass

                if not rep_link:
                    # Último intento: si sólo hay un link en la lista de representados, usarlo
                    links_lista = [l for l in todos_links_texto if len(l) > 4 and l not in ("Salir", "Inicio", "Mis Comprobantes")]
                    print(f"[AFIP] '{representado}' no encontrado. Links disponibles: {todos_links_texto[:20]}")
                    if len(links_lista) == 1:
                        # Hay un único representado — intentar clickearlo
                        for el in mc.locator("a").all():
                            try:
                                if el.inner_text(timeout=300).strip() == links_lista[0]:
                                    rep_link = el
                                    print(f"[AFIP] Único representado disponible: '{links_lista[0][:60]}'")
                                    break
                            except Exception:
                                pass

                if not rep_link:
                    for tipo in tipos:
                        resultados[tipo] = (
                            None,
                            f"No se encontró el representado '{representado}' en AFIP. "
                            f"Verificá el campo 'Representado en AFIP' en la ficha del cliente. "
                            f"Opciones disponibles: {', '.join(todos_links_texto[:8])}"
                        )
                    return resultados

                rep_link.click()
                mc.wait_for_load_state("networkidle", timeout=10000)
                mc.wait_for_timeout(2000)
                print(f"[AFIP] Representado seleccionado. URL: {mc.url}")

            # ── 4. DESCARGAR EMITIDOS Y RECIBIDOS ────────────────────
            for tipo in tipos:
                try:
                    btn_id = "#btnEmitidos" if tipo == "emitidos" else "#btnRecibidos"
                    print(f"[AFIP] Procesando {tipo}...")

                    # Si no estamos en el menú, volver
                    if "setearContribuyente" not in mc.url and "menuPrincipal" not in mc.url and "comprobantes" not in mc.url:
                        mc.goto("https://fes.afip.gob.ar/mcmp/jsp/menuPrincipal.do", timeout=10000)
                        mc.wait_for_load_state("networkidle")
                        mc.wait_for_timeout(2000)

                    mc.click(btn_id)
                    mc.wait_for_load_state("networkidle", timeout=15000)
                    mc.wait_for_timeout(3000)
                    print(f"[AFIP] {tipo} abierto. URL: {mc.url}")

                    # Llenar rango de fechas
                    print(f"[AFIP] Fecha: {rango_fecha}")
                    fecha_el = mc.locator("#fechaEmision")
                    fecha_el.click()
                    mc.wait_for_timeout(500)
                    fecha_el.fill(rango_fecha)
                    mc.keyboard.press("Escape")
                    mc.wait_for_timeout(500)

                    # Buscar
                    mc.click("#buscarComprobantes")
                    mc.wait_for_load_state("networkidle", timeout=20000)
                    mc.wait_for_timeout(4000)

                    # Verificar resultados
                    filas = mc.locator("#tablaDataTables tbody tr").count()
                    print(f"[AFIP] Resultados: {filas} filas")

                    if filas == 0:
                        resultados[tipo] = (None, f"Sin comprobantes {tipo} en {periodo_desde}/{periodo_hasta}")
                        # Volver al menú para el siguiente tipo
                        mc.goto("https://fes.afip.gob.ar/mcmp/jsp/menuPrincipal.do", timeout=10000)
                        mc.wait_for_load_state("networkidle")
                        mc.wait_for_timeout(2000)
                        continue

                    # Descargar CSV
                    print("[AFIP] Descargando CSV...")
                    csv_btn = mc.locator(".sinbor").first
                    if not csv_btn.is_visible(timeout=3000):
                        csv_btn = mc.locator("button:has-text('CSV'), a:has-text('CSV')").first

                    filename = f"{_strip_cuit(cliente_cuit)}_{tipo}_{periodo_desde}_{periodo_hasta}.xlsx"
                    filepath = os.path.join(download_dir, filename)

                    with mc.expect_download(timeout=30000) as dl_info:
                        csv_btn.click()

                    download = dl_info.value

                    # AFIP entrega el CSV dentro de un ZIP — extraer y convertir a xlsx
                    tmp_path = os.path.join(download_dir, f"_tmp_{tipo}.zip")
                    download.save_as(tmp_path)

                    csv_content = None
                    if zipfile.is_zipfile(tmp_path):
                        with zipfile.ZipFile(tmp_path, "r") as zf:
                            csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
                            if csv_names:
                                with zf.open(csv_names[0]) as src:
                                    csv_content = src.read()
                                print(f"[AFIP] ZIP extraído: {csv_names[0]}")
                    else:
                        with open(tmp_path, "rb") as f:
                            csv_content = f.read()

                    if os.path.exists(tmp_path):
                        os.remove(tmp_path)

                    if csv_content:
                        _csv_a_xlsx(csv_content, filepath)
                        print(f"[AFIP] Convertido a xlsx: {filepath}")
                    else:
                        resultados[tipo] = (None, f"No se pudo extraer contenido del ZIP de {tipo}")
                        continue

                    print(f"[AFIP] Guardado: {filepath}")
                    resultados[tipo] = (filepath, None)

                    # Volver al menú para el siguiente tipo
                    if tipo != tipos[-1]:
                        mc.goto("https://fes.afip.gob.ar/mcmp/jsp/menuPrincipal.do", timeout=10000)
                        mc.wait_for_load_state("networkidle")
                        mc.wait_for_timeout(2000)

                except PWTimeout as e:
                    msg = str(e)[:150]
                    print(f"[AFIP] Timeout en {tipo}: {msg}")
                    resultados[tipo] = (None, f"Timeout en {tipo}: {msg}")
                except Exception as e:
                    print(f"[AFIP] Error en {tipo}: {type(e).__name__}: {e}")
                    resultados[tipo] = (None, f"Error en {tipo}: {str(e)[:150]}")

        except PWTimeout as e:
            msg = str(e)[:200]
            print(f"[AFIP] Timeout global: {msg}")
            for tipo in tipos:
                if resultados[tipo][1] == "No ejecutado":
                    resultados[tipo] = (None, f"Timeout: {msg[:120]}")
        except Exception as e:
            print(f"[AFIP] Error global: {type(e).__name__}: {e}")
            for tipo in tipos:
                if resultados[tipo][1] == "No ejecutado":
                    resultados[tipo] = (None, f"Error: {str(e)[:150]}")
        finally:
            browser.close()

    return resultados


async def ejecutar_descarga(
    afip_cuit: str,
    afip_password: str,
    cliente_cuit: str,
    representado: str,
    periodo_desde: str,
    periodo_hasta: str,
    tipos: list,
    storage_base: str,
) -> dict:
    download_dir = os.path.join(
        storage_base,
        _strip_cuit(cliente_cuit),
        f"{periodo_desde}_{periodo_hasta}",
    )
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _executor,
        _descargar_todos_sync,
        afip_cuit,
        afip_password,
        representado,
        periodo_desde,
        periodo_hasta,
        tipos,
        cliente_cuit,
        download_dir,
    )
